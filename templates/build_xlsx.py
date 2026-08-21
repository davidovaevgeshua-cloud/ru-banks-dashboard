"""Build a friendly Excel template for quarterly IFRS reporting input."""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

OUT = "/home/user/workspace/deploy/templates/quarterly_report.xlsx"

# Column groups: (group_title, group_color, [(field, header, unit_hint, width, comment)])
GROUPS = [
    ("Идентификация", "1F4E78", [
        ("ticker",           "Тикер",         "выпадающий список", 12,
         "Обязательно. Один из 7 тикеров дашборда: SBER, VTBR, T, DOMRF, SVCB, BSPB, MBNK."),
        ("bank_name",        "Банк",          "автозаполнение", 20,
         "Заполнится автоматически по тикеру."),
        ("quarter",          "Квартал",       "2026Q2",         10,
         "Обязательно. Формат YYYYQn."),
        ("period_end",       "Конец периода", "2026-06-30",     14,
         "Обязательно. Последний день квартала."),
        ("publication_date", "Дата публикации","2026-08-15",    14,
         "Обязательно. Когда банк выпустил отчёт."),
        ("source_url",       "Источник (URL)","ссылка на пресс-релиз / датабук", 30,
         "Опционально. Ссылка на страницу с отчётом."),
        ("notes",            "Комментарий",   "разовые эффекты, M&A, изменение методологии", 30,
         "Опционально. Что важно помнить об этом квартале."),
    ]),
    ("Квартальные показатели (за квартал, не LTM)", "2E75B6", [
        ("net_income_parent_q", "ЧП на акционеров (кв)", "млрд ₽", 14,
         "Прибыль за квартал, приходящаяся на акционеров банка. Из отчёта о совокупном доходе."),
        ("roe_q_pct",           "ROE (кв)",              "%",     10,
         "Return on Equity за квартал в процентах. Число без %."),
        ("roa_q_pct",           "ROA (кв)",              "%",     10,
         "Return on Assets за квартал."),
        ("nim_q_pct",           "NIM (кв)",              "%",     10,
         "Net Interest Margin за квартал."),
        ("cir_q_pct",           "CIR (кв)",              "%",     10,
         "Cost-to-Income Ratio за квартал."),
        ("cor_q_pct",           "COR (кв)",              "%",     10,
         "Cost of Risk за квартал."),
        ("roe_adj_q_pct",       "ROE скорр. (кв)",       "%",     14,
         "ROE, очищенный от разовых эффектов. Если банк не раскрывает — пусто."),
    ]),
    ("LTM (последние 12 месяцев)", "70AD47", [
        ("net_income_parent_ltm", "ЧП на акционеров LTM", "млрд ₽", 16,
         "Сумма чистой прибыли за 4 последних квартала. Если банк даёт цифру — берём её, иначе я сам сложу 4 квартала."),
        ("nii_ltm",               "ЧПД LTM",              "млрд ₽", 14,
         "Net Interest Income за LTM (чистый процентный доход)."),
        ("nfi_ltm",               "ЧКД LTM",              "млрд ₽", 14,
         "Net Fee Income за LTM (чистый комиссионный доход)."),
        ("roe_ltm_pct",           "ROE LTM",              "%",      12,
         "ROE за последние 12 месяцев."),
    ]),
    ("Баланс на конец периода", "C55A11", [
        ("assets",         "Активы",           "млрд ₽", 14,
         "Активы группы на дату конца периода."),
        ("equity_parent",  "Капитал акционеров","млрд ₽",18,
         "ОБЯЗАТЕЛЬНО для расчёта P/B. Капитал, приходящийся на акционеров банка."),
        ("loans",          "Кредиты (нетто)",  "млрд ₽", 14,
         "Кредитный портфель за вычетом резервов."),
        ("deposits",       "Депозиты",         "млрд ₽", 14,
         "Средства клиентов."),
        ("npl_ratio_pct",  "NPL %",            "%",      10,
         "Доля просроченных кредитов (обычно NPL 90+)."),
        ("car_h20_pct",    "Н20/CAR общ.",     "%",      12,
         "Норматив достаточности общего капитала."),
        ("car_t1_pct",     "CAR Tier 1",       "%",      12,
         "Норматив достаточности капитала первого уровня."),
        ("ni_growth_yoy_pct","Рост ЧП YoY",    "%",      12,
         "Опционально. Прирост LTM-прибыли год-к-году. Если пусто — пересчитаю сам."),
    ]),
    ("Дивиденды (если анонсированы вместе с отчётом)", "7030A0", [
        ("dps_rub",   "DPS", "₽/акция", 10,
         "Рекомендованный дивиденд на акцию. Оставь пусто, если новой рекомендации не было."),
        ("rec_date",  "Дата отсечки", "YYYY-MM-DD", 14,
         "Дата закрытия реестра."),
        ("for_period","За период",    "1H2026 / 2025 / 9M2026", 14,
         "За какой период выплата."),
        ("div_type",  "Тип",          "interim / final", 10,
         "Промежуточный или годовой."),
    ]),
    ("Прогноз (опционально, если менеджмент обновил гайденс)", "808080", [
        ("fwd_2026_ni", "NI 2026e", "млрд ₽", 12, "Прогноз ЧП на 2026. Оставь пусто — сохраню текущий."),
        ("fwd_2026_eq", "EQ 2026e", "млрд ₽", 12, "Прогноз капитала на конец 2026."),
        ("fwd_2027_ni", "NI 2027e", "млрд ₽", 12, ""),
        ("fwd_2027_eq", "EQ 2027e", "млрд ₽", 12, ""),
        ("fwd_2028_ni", "NI 2028e", "млрд ₽", 12, ""),
        ("fwd_2028_eq", "EQ 2028e", "млрд ₽", 12, ""),
    ]),
]

BANKS = [
    ("SBER",  "Сбербанк"),
    ("VTBR",  "ВТБ"),
    ("T",     "Т-Технологии"),
    ("DOMRF", "ДОМ.РФ"),
    ("SVCB",  "Совкомбанк"),
    ("BSPB",  "Банк Санкт-Петербург"),
    ("MBNK",  "МТС Банк"),
]

# ---------- workbook ----------
wb = Workbook()

# ============ Sheet 1: Отчёт ============
ws = wb.active
ws.title = "Отчёт"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C4"

thin = Side(style="thin", color="D9D9D9")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

# Row 1: group headers (merged); Row 2: field header; Row 3: unit hint
col = 1
field_positions = {}  # field_id -> column index
for group_title, color, fields in GROUPS:
    start = col
    for field_id, header, unit_hint, width, comment in fields:
        cell_group = ws.cell(row=1, column=col, value=group_title)
        cell_group.fill = PatternFill("solid", fgColor=color)
        cell_group.font = Font(bold=True, color="FFFFFF", size=11)
        cell_group.alignment = Alignment(horizontal="center", vertical="center")

        cell_hdr = ws.cell(row=2, column=col, value=header)
        cell_hdr.fill = PatternFill("solid", fgColor="F2F2F2")
        cell_hdr.font = Font(bold=True, size=10)
        cell_hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_hdr.border = border_all
        if comment:
            cell_hdr.comment = Comment(comment, "template")

        cell_unit = ws.cell(row=3, column=col, value=unit_hint)
        cell_unit.fill = PatternFill("solid", fgColor="FAFAFA")
        cell_unit.font = Font(italic=True, size=9, color="808080")
        cell_unit.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_unit.border = border_all

        ws.column_dimensions[get_column_letter(col)].width = width
        field_positions[field_id] = col
        col += 1
    end = col - 1
    if end > start:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 30
ws.row_dimensions[3].height = 18

# Data rows: one per bank, prefilled ticker + bank name
NUM_DATA_ROWS = 7
tk_col = field_positions["ticker"]
bank_col = field_positions["bank_name"]

for i, (tk, name) in enumerate(BANKS):
    r = 4 + i
    ws.cell(row=r, column=tk_col, value=tk).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=bank_col, value=name).alignment = Alignment(horizontal="left")
    # light stripe
    if i % 2 == 1:
        for c in range(1, col):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F7F9FC")

    # style all cells in row
    for c in range(1, col):
        cell = ws.cell(row=r, column=c)
        cell.border = border_all
        # numeric fields — right align + number format
        fid = next((f for f, cc in field_positions.items() if cc == c), None)
        if fid and fid not in ("ticker","bank_name","quarter","period_end","publication_date",
                               "source_url","notes","rec_date","for_period","div_type"):
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0.00"
        elif fid in ("period_end","publication_date","rec_date"):
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "yyyy-mm-dd"

    ws.row_dimensions[r].height = 20

# Data validation
# Ticker dropdown
dv_ticker = DataValidation(type="list",
                           formula1='"SBER,VTBR,T,DOMRF,SVCB,BSPB,MBNK"',
                           allow_blank=True)
dv_ticker.error = "Только один из тикеров дашборда"
dv_ticker.errorTitle = "Неверный тикер"
ws.add_data_validation(dv_ticker)
dv_ticker.add(f"{get_column_letter(tk_col)}4:{get_column_letter(tk_col)}{3+NUM_DATA_ROWS}")

# Quarter format hint
dv_quarter = DataValidation(type="textLength", operator="equal", formula1="6", allow_blank=True)
dv_quarter.error = "Формат YYYYQn, например 2026Q2 (6 символов)"
dv_quarter.errorTitle = "Формат квартала"
q_col = field_positions["quarter"]
ws.add_data_validation(dv_quarter)
dv_quarter.add(f"{get_column_letter(q_col)}4:{get_column_letter(q_col)}{3+NUM_DATA_ROWS}")

# Div type dropdown
dv_type = DataValidation(type="list", formula1='"interim,final"', allow_blank=True)
dt_col = field_positions["div_type"]
ws.add_data_validation(dv_type)
dv_type.add(f"{get_column_letter(dt_col)}4:{get_column_letter(dt_col)}{3+NUM_DATA_ROWS}")

# Highlight the "must-fill" cells (light yellow border tint via fill)
MUST = ["ticker","quarter","period_end","publication_date",
        "net_income_parent_ltm","equity_parent"]
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
for fid in MUST:
    c = field_positions[fid]
    for i in range(NUM_DATA_ROWS):
        cell = ws.cell(row=4+i, column=c)
        # only tint the empty ones (not ticker/bank prefilled)
        if fid in ("ticker",):
            continue
        if cell.value in (None, ""):
            cell.fill = yellow_fill

# ============ Sheet 2: Инструкция ============
ws2 = wb.create_sheet("Инструкция")
ws2.sheet_view.showGridLines = False

instructions = [
    ("Как заполнять", 14, True, "1F4E78"),
    ("", 11, False, None),
    ("• Одна строка = один банк, один квартал", 11, False, None),
    ("• Заполняй только те банки, которые действительно отчитались. Пустые строки я проигнорирую.", 11, False, None),
    ("• Жёлтые ячейки — обязательный минимум:", 11, False, None),
    ("    – Тикер, Квартал, Конец периода, Дата публикации", 11, False, None),
    ("    – ЧП на акционеров LTM (или квартальная, я досчитаю)", 11, False, None),
    ("    – Капитал акционеров (нужен для P/B)", 11, False, None),
    ("• Остальное — опционально, но чем больше заполнено, тем полнее обновление.", 11, False, None),
    ("", 11, False, None),
    ("Единицы измерения", 14, True, "1F4E78"),
    ("", 11, False, None),
    ("• Все денежные величины — в млрд рублей (например, 55.87 = 55.87 млрд ₽)", 11, False, None),
    ("• Все проценты — числом БЕЗ знака %: 22.5, а не 22.5% и не 0.225", 11, False, None),
    ("• Даты — в формате YYYY-MM-DD", 11, False, None),
    ("", 11, False, None),
    ("Что произойдёт после того, как ты пришлёшь заполненный файл", 14, True, "1F4E78"),
    ("", 11, False, None),
    ("1. Я обновлю квартальные ряды в payload дашборда (NIM, CIR, COR, ROE, NPL, CAR, equity, assets и т.д.)", 11, False, None),
    ("2. Пересчитаю LTM-мультипликаторы (P/E, P/B, P/TBV, ROE) от новой прибыли/капитала и текущей цены", 11, False, None),
    ("3. Обновлю прогнозные NI/EQ, если ты обновила блок «Прогноз»; иначе оставлю прежние", 11, False, None),
    ("4. Если есть новая рекомендация по дивидендам — добавлю в форвардную таблицу и таблицу выплат", 11, False, None),
    ("5. Установлю last[TK].period на дату конца периода", 11, False, None),
    ("6. Обновлю state/reports_last_seen.json: seen[TK] = новый квартал (чтобы уведомления про этот же квартал больше не приходили)", 11, False, None),
    ("7. Закоммичу — GitHub Pages автоматически развернёт обновление", 11, False, None),
    ("", 11, False, None),
    ("Если что-то непонятно", 14, True, "1F4E78"),
    ("", 11, False, None),
    ("Можно просто прислать датабук/пресс-релиз (PDF/Excel) и написать «извлеки цифры» — я заполню шаблон сам и покажу тебе на подтверждение.", 11, False, None),
]

ws2.column_dimensions["A"].width = 110
for i, (text, size, bold, color) in enumerate(instructions, start=1):
    c = ws2.cell(row=i, column=1, value=text)
    c.font = Font(size=size, bold=bold, color=color or "000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")

# Page setup — landscape + fit to width
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A3
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
ws.print_title_rows = '1:3'

ws2.page_setup.orientation = ws2.ORIENTATION_PORTRAIT
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUT)
print(f"Saved: {OUT}")
